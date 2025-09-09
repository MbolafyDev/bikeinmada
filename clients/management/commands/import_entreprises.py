import datetime
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from clients.models import Entreprise

class Command(BaseCommand):
    help = 'Importe les entreprises depuis un fichier Excel'

    def add_arguments(self, parser):
        parser.add_argument('fichier_excel', type=str, help='Chemin vers le fichier Excel (ex: entreprises.xlsx)')

    def handle(self, *args, **kwargs):
        fichier_excel = kwargs['fichier_excel']
        self.stdout.write("🔧 Démarrage de l'importation")

        try:
            wb = load_workbook(fichier_excel)
            ws = wb.active
        except Exception as e:
            self.stderr.write(f"❌ Erreur d'ouverture : {e}")
            return

        colonnes = [str(cell.value).strip() for cell in ws[1] if cell.value]

        def parse_date(val):
            if isinstance(val, datetime.datetime):
                return val.date()
            if isinstance(val, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                    try:
                        return datetime.datetime.strptime(val, fmt).date()
                    except ValueError:
                        continue
            return None

        nb_imports = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(colonnes, row[:len(colonnes)]))
            data['date_debut'] = parse_date(data.get('date_debut'))
            data['date_cin'] = parse_date(data.get('date_cin'))

            try:
                entreprise = Entreprise(**data)
                entreprise.save()
                self.stdout.write(f"✅ Importé : {entreprise.raison_sociale}")
                nb_imports += 1
            except Exception as e:
                self.stderr.write(f"❌ Erreur sur {data.get('raison_sociale', 'Inconnu')} : {e}")

        self.stdout.write(f"✅ Importation terminée ({nb_imports} entrées)")
